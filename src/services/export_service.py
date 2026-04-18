"""
Export Service - generates Excel, CSV, and JSON reports from BOQ data.

Enhanced with:
- Section subtotals and grand total
- Audit trail sheet (traceability: base qty, waste %, source elements)
- Unit-aware number formatting (2 dec for areas, 3 for volumes, 0 for counts)
- Currency column support
"""

from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.translations.strings import get_export_labels
from src.utils.logger import get_logger

logger = get_logger("export_service")


# Styling constants
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Calibri", size=11, bold=True, color="2F5496")
ITEM_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", size=12, color="666666")
TOTAL_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FONT = Font(name="Calibri", size=11, bold=True)
GRAND_TOTAL_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
GRAND_TOTAL_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CONFIDENCE_FILLS = {
    "high": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "medium": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "low": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}
CONFIDENCE_FONTS = {
    "high": Font(name="Calibri", size=9, color="006100"),
    "medium": Font(name="Calibri", size=9, color="9C5700"),
    "low": Font(name="Calibri", size=9, bold=True, color="9C0006"),
}


def _get_number_format(unit: str) -> str:
    """Return Excel number format appropriate for the unit type."""
    unit_lower = unit.lower().strip()
    if unit_lower in ("nr", "set", "pcs"):
        return "#,##0"
    elif unit_lower == "m3":
        return "#,##0.000"
    elif unit_lower == "kg":
        return "#,##0.0"
    else:
        # m2, m, litre, and default
        return "#,##0.00"


class ExportService:
    """Generates reports from BOQ data."""

    def export_excel(
        self, boq_data: dict[str, Any], output_path: str | Path,
        language: str = "en", currency: str = "",
    ) -> Path:
        """Export BOQ data to a professional Excel spreadsheet.

        Args:
            boq_data: The BOQ data dict from BOQGeneratorAgent.
            output_path: Where to save the .xlsx file.
            language: Language for labels.
            currency: Optional currency symbol for rate/amount columns.

        Returns:
            Path to the generated file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        labels = get_export_labels(language)
        is_rtl = language == "ar"

        wb = Workbook()

        # --- Sheet 1: BOQ (main report) ---
        ws = wb.active
        ws.title = labels["sheet_boq"]
        if is_rtl:
            ws.sheet_properties.isRightToLeft = True
        self._write_boq_sheet(ws, boq_data, labels, currency)

        # --- Sheet 2: Material Summary ---
        ws2 = wb.create_sheet(labels["sheet_summary"])
        if is_rtl:
            ws2.sheet_properties.isRightToLeft = True
        self._write_summary_sheet(ws2, boq_data, labels)

        # --- Sheet 3: Audit Trail ---
        ws3 = wb.create_sheet(labels.get("sheet_audit", "Audit Trail"))
        if is_rtl:
            ws3.sheet_properties.isRightToLeft = True
        self._write_audit_sheet(ws3, boq_data, labels)

        wb.save(str(output_path))
        logger.info(f"Excel report saved: {output_path}")
        return output_path

    def _write_boq_sheet(
        self, ws, boq_data: dict[str, Any],
        labels: dict[str, str], currency: str = "",
    ) -> None:
        """Write the main BOQ sheet with subtotals and grand total."""
        # Column widths
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 12

        row = 1

        # Title
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=labels["boq_title"])
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal="center")
        row += 1

        # Project info
        project_name = boq_data.get("project_name", "Untitled")
        building_name = boq_data.get("building_name", "")
        date_str = datetime.now().strftime("%Y-%m-%d")

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=f"{labels['project']}: {project_name}")
        cell.font = SUBTITLE_FONT
        cell.alignment = Alignment(horizontal="center")
        row += 1

        if building_name:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            cell = ws.cell(row=row, column=1, value=f"{labels['building']}: {building_name}")
            cell.font = SUBTITLE_FONT
            cell.alignment = Alignment(horizontal="center")
            row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(
            row=row, column=1,
            value=f"{labels['date']}: {date_str}  |  {labels['prepared_by']}: {boq_data.get('prepared_by', 'Metraj AI')}",
        )
        cell.font = Font(name="Calibri", size=9, color="999999")
        cell.alignment = Alignment(horizontal="center")
        row += 2

        # Column headers
        rate_label = f"{labels['rate']}" + (f" ({currency})" if currency else "")
        amount_label = f"{labels['amount']}" + (f" ({currency})" if currency else "")
        headers = [
            labels["item_no"], labels["description"], labels["unit"],
            labels["quantity"], rate_label, amount_label, "Confidence",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        # Track first data row for grand total formula
        data_start_row = row
        amount_col_letter = get_column_letter(6)  # Column F

        # Sections and items
        section_subtotal_rows: list[int] = []

        for section in boq_data.get("sections", []):
            # Section header row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            cell = ws.cell(
                row=row, column=1,
                value=f"{section['section_no']}. {section['title']}",
            )
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            cell.border = THIN_BORDER
            for col in range(2, 8):
                ws.cell(row=row, column=col).fill = SECTION_FILL
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

            section_item_rows: list[int] = []

            # Items
            for item in section.get("items", []):
                ws.cell(row=row, column=1, value=item["item_no"]).font = ITEM_FONT
                ws.cell(row=row, column=2, value=item["description"]).font = ITEM_FONT
                ws.cell(row=row, column=3, value=item["unit"]).font = ITEM_FONT

                # Unit-aware number formatting
                num_fmt = _get_number_format(item["unit"])

                qty_cell = ws.cell(row=row, column=4, value=item["quantity"])
                qty_cell.font = ITEM_FONT
                qty_cell.number_format = num_fmt

                rate_cell = ws.cell(row=row, column=5, value=item.get("rate"))
                rate_cell.font = ITEM_FONT
                rate_cell.number_format = "#,##0.00"

                amount_cell = ws.cell(row=row, column=6, value=item.get("amount"))
                amount_cell.font = ITEM_FONT
                amount_cell.number_format = "#,##0.00"

                # Confidence column (G)
                conf = item.get("confidence", {})
                conf_level = conf.get("level", "medium") if conf else "medium"
                conf_cell = ws.cell(row=row, column=7, value=conf_level.upper())
                conf_cell.font = CONFIDENCE_FONTS.get(conf_level, ITEM_FONT)
                conf_cell.fill = CONFIDENCE_FILLS.get(conf_level, PatternFill())
                conf_cell.alignment = Alignment(horizontal="center", vertical="center")

                # Borders and alignment
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = THIN_BORDER
                    if col < 7:
                        ws.cell(row=row, column=col).alignment = Alignment(
                            horizontal="left" if col == 2 else "center",
                            vertical="center",
                        )

                section_item_rows.append(row)
                row += 1

            # Section subtotal row
            if section_item_rows:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                subtotal_label_cell = ws.cell(
                    row=row, column=1,
                    value=f"    {labels.get('subtotal', 'Subtotal')} - {section['title']}",
                )
                subtotal_label_cell.font = TOTAL_FONT
                subtotal_label_cell.fill = TOTAL_FILL
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = TOTAL_FILL
                    ws.cell(row=row, column=col).border = THIN_BORDER

                # SUM formula for the amount column
                first = section_item_rows[0]
                last = section_item_rows[-1]
                sum_cell = ws.cell(
                    row=row, column=6,
                    value=f"=SUM({amount_col_letter}{first}:{amount_col_letter}{last})",
                )
                sum_cell.font = TOTAL_FONT
                sum_cell.fill = TOTAL_FILL
                sum_cell.border = THIN_BORDER
                sum_cell.number_format = "#,##0.00"

                section_subtotal_rows.append(row)
                row += 1

            row += 1  # gap between sections

        # Grand total row
        if section_subtotal_rows:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            gt_label = ws.cell(
                row=row, column=1,
                value=labels.get("grand_total", "GRAND TOTAL"),
            )
            gt_label.font = GRAND_TOTAL_FONT
            gt_label.fill = GRAND_TOTAL_FILL
            gt_label.alignment = Alignment(horizontal="right")
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = GRAND_TOTAL_FILL
                ws.cell(row=row, column=col).border = THIN_BORDER

            # SUM of subtotals
            refs = "+".join(f"{amount_col_letter}{r}" for r in section_subtotal_rows)
            gt_cell = ws.cell(row=row, column=6, value=f"={refs}")
            gt_cell.font = GRAND_TOTAL_FONT
            gt_cell.fill = GRAND_TOTAL_FILL
            gt_cell.border = THIN_BORDER
            gt_cell.number_format = "#,##0.00"
            row += 2

        # Footer
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=labels["footer"])
        cell.font = Font(name="Calibri", size=8, italic=True, color="999999")
        cell.alignment = Alignment(horizontal="center")

    def _write_summary_sheet(
        self, ws, boq_data: dict[str, Any],
        labels: dict[str, str],
    ) -> None:
        """Write a material summary sheet (aggregated by material type)."""
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 20

        row = 1

        # Title
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value=labels["material_summary"])
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal="center")
        row += 2

        # Headers
        headers = [
            labels["no"], labels["material"], labels["unit"],
            labels["total_qty"], labels["category"],
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Items
        item_no = 0
        for section in boq_data.get("sections", []):
            for item in section.get("items", []):
                item_no += 1
                ws.cell(row=row, column=1, value=item_no).border = THIN_BORDER
                ws.cell(row=row, column=2, value=item["description"]).border = THIN_BORDER

                unit_cell = ws.cell(row=row, column=3, value=item["unit"])
                unit_cell.border = THIN_BORDER
                unit_cell.alignment = Alignment(horizontal="center")

                qty_cell = ws.cell(row=row, column=4, value=item["quantity"])
                qty_cell.border = THIN_BORDER
                qty_cell.number_format = _get_number_format(item["unit"])
                qty_cell.alignment = Alignment(horizontal="right")

                cat_cell = ws.cell(row=row, column=5, value=section["title"])
                cat_cell.border = THIN_BORDER
                cat_cell.font = Font(name="Calibri", size=9, color="666666")

                row += 1

    def _write_audit_sheet(
        self, ws, boq_data: dict[str, Any],
        labels: dict[str, str],
    ) -> None:
        """Write an audit trail sheet showing base quantities, waste factors,
        and source element counts for each BOQ line item."""
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 15

        row = 1

        # Title
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(
            row=row, column=1,
            value=labels.get("audit_title", "QUANTITY AUDIT TRAIL"),
        )
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal="center")
        row += 2

        # Headers
        headers = [
            labels["item_no"],
            labels["description"],
            labels["unit"],
            labels.get("base_qty", "Base Qty"),
            labels.get("waste_pct", "Waste %"),
            labels.get("total_qty_with_waste", "Total Qty (incl. waste)"),
            labels.get("source_elements_count", "Source Elements"),
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1

        # Data rows
        for section in boq_data.get("sections", []):
            # Section header
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            cell = ws.cell(
                row=row, column=1,
                value=f"{section['section_no']}. {section['title']}",
            )
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = SECTION_FILL
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

            for item in section.get("items", []):
                num_fmt = _get_number_format(item["unit"])

                ws.cell(row=row, column=1, value=item["item_no"]).border = THIN_BORDER
                ws.cell(row=row, column=2, value=item["description"]).border = THIN_BORDER
                ws.cell(row=row, column=3, value=item["unit"]).border = THIN_BORDER

                base_cell = ws.cell(
                    row=row, column=4,
                    value=item.get("base_quantity", item["quantity"]),
                )
                base_cell.border = THIN_BORDER
                base_cell.number_format = num_fmt

                waste_pct = item.get("waste_factor", 0)
                waste_cell = ws.cell(row=row, column=5, value=waste_pct)
                waste_cell.border = THIN_BORDER
                waste_cell.number_format = "0.0%"

                total_cell = ws.cell(row=row, column=6, value=item["quantity"])
                total_cell.border = THIN_BORDER
                total_cell.number_format = num_fmt

                elem_count = item.get("element_count", 0)
                count_cell = ws.cell(row=row, column=7, value=elem_count)
                count_cell.border = THIN_BORDER
                count_cell.alignment = Alignment(horizontal="center")

                for col in range(1, 8):
                    ws.cell(row=row, column=col).font = ITEM_FONT

                row += 1

    def export_csv(
        self, boq_data: dict[str, Any], output_path: str | Path,
        language: str = "en",
    ) -> Path:
        """Export BOQ data to a CSV file."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        labels = get_export_labels(language)

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                labels["item_no"], labels["description"], labels["unit"],
                labels["quantity"],
                labels.get("base_qty", "Base Qty"),
                labels.get("waste_pct", "Waste %"),
                labels["rate"], labels["amount"],
                labels["section"],
            ])

            for section in boq_data.get("sections", []):
                for item in section.get("items", []):
                    writer.writerow([
                        item["item_no"],
                        item["description"],
                        item["unit"],
                        item["quantity"],
                        item.get("base_quantity", ""),
                        f"{item.get('waste_factor', 0):.1%}" if item.get("waste_factor") else "",
                        item.get("rate", ""),
                        item.get("amount", ""),
                        section["title"],
                    ])

        logger.info(f"CSV report saved: {output_path}")
        return output_path

    def export_json(
        self, boq_data: dict[str, Any], output_path: str | Path
    ) -> Path:
        """Export BOQ data to a JSON file."""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(boq_data, f, indent=2, ensure_ascii=False)

        logger.info(f"JSON report saved: {output_path}")
        return output_path
