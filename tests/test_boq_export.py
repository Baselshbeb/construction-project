"""
Tests for BOQGeneratorAgent and ExportService - verifies BOQ structure,
section ordering, item numbering, and export formats.
"""

import json
import csv
import tempfile
import pytest
from pathlib import Path

from src.agents.boq_generator import BOQGeneratorAgent
from src.services.export_service import ExportService


@pytest.fixture
def boq_gen():
    return BOQGeneratorAgent()


@pytest.fixture
def export_svc():
    return ExportService()


def make_material(description, unit, total_quantity, category):
    return {
        "description": description,
        "unit": unit,
        "quantity": total_quantity * 0.95,
        "total_quantity": total_quantity,
        "waste_factor": 0.05,
        "category": category,
        "source_elements": [1],
    }


def make_state(materials=None, building_info=None):
    return {
        "material_list": materials or [],
        "building_info": building_info or {},
        "processing_log": [],
    }


# ---- BOQ Generator Tests ----

class TestBOQGenerator:
    @pytest.mark.asyncio
    async def test_empty_materials_no_crash(self, boq_gen):
        state = make_state(materials=[])
        result = await boq_gen.execute(state)
        assert result.get("boq_data") is None

    @pytest.mark.asyncio
    async def test_sections_follow_config_order(self, boq_gen):
        """Sections should follow SECTION_CONFIG order, not insertion order."""
        state = make_state(materials=[
            make_material("Window unit", "nr", 10, "windows"),
            make_material("Concrete C25/30", "m3", 50, "frame"),
            make_material("Bricks", "nr", 1000, "internal_walls"),
        ])
        result = await boq_gen.execute(state)
        boq = result["boq_data"]

        # frame comes before internal_walls, which comes before windows
        titles = [s["title"] for s in boq["sections"]]
        assert titles.index("Structural Frame (Columns & Beams)") < titles.index("Internal Walls & Partitions")
        assert titles.index("Internal Walls & Partitions") < titles.index("Windows")

    @pytest.mark.asyncio
    async def test_empty_categories_skipped(self, boq_gen):
        """Categories with no materials don't appear as sections."""
        state = make_state(materials=[
            make_material("Concrete", "m3", 50, "frame"),
        ])
        result = await boq_gen.execute(state)
        boq = result["boq_data"]
        assert len(boq["sections"]) == 1
        assert boq["sections"][0]["title"] == "Structural Frame (Columns & Beams)"

    @pytest.mark.asyncio
    async def test_item_numbering_format(self, boq_gen):
        """Item numbers should be section.item with zero-padded item index."""
        state = make_state(materials=[
            make_material("Concrete", "m3", 50, "frame"),
            make_material("Steel", "kg", 5000, "frame"),
            make_material("Paint", "m2", 100, "internal_walls"),
        ])
        result = await boq_gen.execute(state)
        boq = result["boq_data"]

        # Frame section = section 1
        frame_items = boq["sections"][0]["items"]
        assert frame_items[0]["item_no"] == "1.01"
        assert frame_items[1]["item_no"] == "1.02"

        # Internal walls = section 2
        walls_items = boq["sections"][1]["items"]
        assert walls_items[0]["item_no"] == "2.01"

    @pytest.mark.asyncio
    async def test_uncategorized_materials_in_other(self, boq_gen):
        """Materials with no category go to 'Other Items' section."""
        state = make_state(materials=[
            make_material("Something", "nr", 5, None),
        ])
        result = await boq_gen.execute(state)
        boq = result["boq_data"]
        assert boq["sections"][-1]["title"] == "Other Items"

    @pytest.mark.asyncio
    async def test_project_name_from_building_info(self, boq_gen):
        state = make_state(
            materials=[make_material("Concrete", "m3", 50, "frame")],
            building_info={"project_name": "Test Tower", "building_name": "Block A"},
        )
        result = await boq_gen.execute(state)
        boq = result["boq_data"]
        assert boq["project_name"] == "Test Tower"
        assert boq["building_name"] == "Block A"

    @pytest.mark.asyncio
    async def test_default_project_name(self, boq_gen):
        """Missing or empty project name falls back to 'Untitled Project'."""
        state = make_state(
            materials=[make_material("Concrete", "m3", 50, "frame")],
            building_info={"project_name": ""},
        )
        result = await boq_gen.execute(state)
        assert result["boq_data"]["project_name"] == "Untitled Project"

    @pytest.mark.asyncio
    async def test_total_counts(self, boq_gen):
        state = make_state(materials=[
            make_material("Concrete", "m3", 50, "frame"),
            make_material("Steel", "kg", 5000, "frame"),
            make_material("Bricks", "nr", 1000, "internal_walls"),
        ])
        result = await boq_gen.execute(state)
        boq = result["boq_data"]
        assert boq["total_sections"] == 2
        assert boq["total_line_items"] == 3


# ---- Export Service Tests ----

class TestExcelExport:
    def test_excel_file_created(self, export_svc):
        """Excel file is created at the specified path."""
        boq_data = {
            "project_name": "Test",
            "building_name": None,
            "prepared_by": "Test",
            "sections": [{
                "section_no": 1,
                "title": "Frame",
                "items": [{
                    "item_no": "1.01",
                    "description": "Concrete",
                    "unit": "m3",
                    "quantity": 50.0,
                    "rate": None,
                    "amount": None,
                }],
            }],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_svc.export_excel(boq_data, Path(tmpdir) / "test.xlsx")
            assert path.exists()
            assert path.suffix == ".xlsx"

    def test_excel_has_two_sheets(self, export_svc):
        """Excel workbook has BOQ and Material Summary sheets."""
        from openpyxl import load_workbook

        boq_data = {
            "project_name": "Test",
            "building_name": "Tower A",
            "prepared_by": "AI",
            "sections": [{
                "section_no": 1,
                "title": "Frame",
                "items": [{
                    "item_no": "1.01",
                    "description": "Concrete",
                    "unit": "m3",
                    "quantity": 50.0,
                    "rate": None,
                    "amount": None,
                }],
            }],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_svc.export_excel(boq_data, Path(tmpdir) / "test.xlsx")
            wb = load_workbook(str(path))
            assert "Bill of Quantities" in wb.sheetnames
            assert "Material Summary" in wb.sheetnames


class TestCSVExport:
    def test_csv_file_created(self, export_svc):
        boq_data = {
            "sections": [{
                "title": "Frame",
                "items": [{
                    "item_no": "1.01",
                    "description": "Concrete",
                    "unit": "m3",
                    "quantity": 50.0,
                    "rate": None,
                    "amount": None,
                }],
            }],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_svc.export_csv(boq_data, Path(tmpdir) / "test.csv")
            assert path.exists()

    def test_csv_contains_correct_data(self, export_svc):
        """CSV has headers and item data."""
        boq_data = {
            "sections": [{
                "title": "Frame",
                "items": [
                    {"item_no": "1.01", "description": "Concrete", "unit": "m3",
                     "quantity": 50.0, "rate": None, "amount": None},
                    {"item_no": "1.02", "description": "Steel", "unit": "kg",
                     "quantity": 5000.0, "rate": 2.5, "amount": 12500.0},
                ],
            }],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_svc.export_csv(boq_data, Path(tmpdir) / "test.csv")
            with open(path, "r", encoding="utf-8") as f:
                reader = list(csv.reader(f))

            # Header row
            assert reader[0][0] == "Item No."
            assert reader[0][-1] == "Section"
            # Data rows
            assert reader[1][0] == "1.01"
            assert reader[1][1] == "Concrete"
            assert reader[1][-1] == "Frame"
            # Rate/amount with None -> empty string
            assert reader[1][4] == ""  # rate
            assert reader[1][5] == ""  # amount

    def test_csv_rate_none_becomes_empty(self, export_svc):
        """None rate/amount should be empty string in CSV."""
        boq_data = {
            "sections": [{
                "title": "T",
                "items": [{"item_no": "1", "description": "X", "unit": "m",
                           "quantity": 1, "rate": None, "amount": None}],
            }],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_svc.export_csv(boq_data, Path(tmpdir) / "test.csv")
            with open(path, "r", encoding="utf-8") as f:
                rows = list(csv.reader(f))
            assert rows[1][4] == ""
            assert rows[1][5] == ""


class TestJSONExport:
    def test_json_file_created(self, export_svc):
        boq_data = {"project_name": "Test", "sections": []}
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_svc.export_json(boq_data, Path(tmpdir) / "test.json")
            assert path.exists()

    def test_json_roundtrip(self, export_svc):
        """JSON should be loadable and match original data."""
        boq_data = {
            "project_name": "Test Project",
            "sections": [{"section_no": 1, "title": "Frame", "items": []}],
        }
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_svc.export_json(boq_data, Path(tmpdir) / "test.json")
            with open(path, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            assert loaded["project_name"] == "Test Project"
            assert len(loaded["sections"]) == 1


class TestExportDirectoryCreation:
    def test_creates_nested_directories(self, export_svc):
        """Export should create parent directories if they don't exist."""
        boq_data = {"project_name": "Test", "sections": []}
        with tempfile.TemporaryDirectory() as tmpdir:
            deep_path = Path(tmpdir) / "a" / "b" / "c" / "test.json"
            path = export_svc.export_json(boq_data, deep_path)
            assert path.exists()
